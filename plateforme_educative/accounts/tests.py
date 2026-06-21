from django.test import TestCase, override_settings
from django.utils import timezone
from django.urls import reverse
from django.contrib.auth import get_user_model
from datetime import timedelta
from django.core import mail

Utilisateur = get_user_model()
from accounts.models import Notification

@override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
class TemporaryPasswordTests(TestCase):
    def setUp(self):
        self.email = "test_user@example.com"
        self.password = "OriginalPassword123!"
        self.user = Utilisateur.objects.create_user(
            email=self.email,
            password=self.password,
            role="ELEVE",
            statut_compte="ACTIVE",
            is_active=True
        )

    def test_user_model_fields(self):
        """Test that default values for temporary password fields are set correctly."""
        self.assertFalse(self.user.is_temp_password)
        self.assertIsNone(self.user.temp_password_created_at)

    def test_custom_password_reset_view(self):
        """Test requesting a password reset generates a temp password and sends an email."""
        response = self.client.post(reverse('accounts:password_reset'), {'email': self.email})
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('accounts:password_reset_done'))

        # Check mail was sent
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn("Votre mot de passe temporaire", mail.outbox[0].subject)
        self.assertIn(self.email, mail.outbox[0].to)

        # Check user fields updated
        self.user.refresh_from_db()
        self.assertTrue(self.user.is_temp_password)
        self.assertIsNotNone(self.user.temp_password_created_at)

    def test_login_with_valid_temp_password(self):
        """Test logging in with a temporary password within the 10-minute window."""
        # Setup temporary password
        temp_pass = "TEMP1234"
        self.user.set_password(temp_pass)
        self.user.is_temp_password = True
        self.user.temp_password_created_at = timezone.now() - timedelta(minutes=5)
        self.user.save()

        # Try to log in
        response = self.client.post(reverse('accounts:login'), {
            'username': self.email,
            'password': temp_pass
        })
        
        # Should redirect to change password view
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('accounts:change_password'))

        # Verify user is logged in
        self.assertIn('_auth_user_id', self.client.session)

    def test_login_with_expired_temp_password(self):
        """Test logging in with a temporary password after the 10-minute window."""
        temp_pass = "TEMP5678"
        self.user.set_password(temp_pass)
        self.user.is_temp_password = True
        self.user.temp_password_created_at = timezone.now() - timedelta(minutes=11)
        self.user.save()

        # Try to log in
        response = self.client.post(reverse('accounts:login'), {
            'username': self.email,
            'password': temp_pass
        })

        # Should render the login page with errors
        self.assertEqual(response.status_code, 200)
        self.assertNotIn('_auth_user_id', self.client.session)
        self.assertFormError(response, 'form', None, "Ce mot de passe temporaire a expiré (validité de 10 minutes). Veuillez effectuer une nouvelle demande.")

    def test_change_password_clears_flags(self):
        """Test that changing password clears the temporary password flags."""
        # Log the user in
        self.client.login(email=self.email, password=self.password)

        # Simulate the user changing their password
        response = self.client.post(reverse('accounts:change_password'), {
            'old_password': self.password,
            'new_password1': 'NewSecurePassword123!',
            'new_password2': 'NewSecurePassword123!'
        })

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('accounts:profile'))

        # Check DB updates
        self.user.refresh_from_db()
        self.assertFalse(self.user.is_temp_password)
        self.assertIsNone(self.user.temp_password_created_at)
        self.assertTrue(self.user.check_password('NewSecurePassword123!'))


class HTMXHistoryRestoreMiddlewareTests(TestCase):
    def test_middleware_removes_hx_request_on_restore_request(self):
        """Test that HTMXHistoryRestoreMiddleware removes the HTTP_HX_REQUEST header when HTTP_HX_HISTORY_RESTORE_REQUEST is true."""
        headers = {
            'HTTP_HX_REQUEST': 'true',
            'HTTP_HX_HISTORY_RESTORE_REQUEST': 'true'
        }
        # Requesting registration page with both headers should bypass partial rendering and return full page.
        response = self.client.get(reverse('accounts:register'), **headers)
        
        # Ensure it contains the main DOCTYPE showing it's a full page rendering
        self.assertContains(response, "<!DOCTYPE html>")

    def test_middleware_adds_vary_hx_request_header(self):
        """Test that HTMXHistoryRestoreMiddleware adds 'HX-Request' to the 'Vary' header of the response."""
        response = self.client.get(reverse('accounts:login'))
        self.assertIn('Vary', response)
        self.assertIn('HX-Request', response['Vary'])

    def test_middleware_disables_cache_on_htmx_requests(self):
        """Test that HTMXHistoryRestoreMiddleware disables caching on HTMX requests."""
        headers = {
            'HTTP_HX_REQUEST': 'true'
        }
        response = self.client.get(reverse('accounts:register'), **headers)
        self.assertEqual(response['Cache-Control'], 'no-cache, no-store, must-revalidate')
        self.assertEqual(response['Pragma'], 'no-cache')
        self.assertEqual(response['Expires'], '0')


class RegistrationNotificationTests(TestCase):
    def setUp(self):
        self.admin = Utilisateur.objects.create_user(
            email="admin_test_notif@example.com",
            password="AdminPassword123!",
            role="ADMIN",
            statut_compte="ACTIVE",
            is_active=True
        )
        self.superuser = Utilisateur.objects.create_superuser(
            email="superuser_test_notif@example.com",
            password="SuperuserPassword123!",
            is_active=True
        )

    def test_student_registration_creates_notifications_for_admins(self):
        """Test that registering a new student generates notification alerts for all admins and superusers."""
        response = self.client.post(reverse('accounts:register'), {
            'email': 'new_student_test@example.com',
            'password1': 'NewStudentPass123!',
            'password2': 'NewStudentPass123!',
            'role': 'ELEVE'
        })
        self.assertEqual(response.status_code, 302)

        # Check notifications generated for admin
        admin_notifs = Notification.objects.filter(destinataire=self.admin)
        self.assertEqual(admin_notifs.count(), 1)
        self.assertEqual(admin_notifs.first().type, 'NOUVELLE_INSCRIPTION')
        self.assertIn('new_student_test@example.com', admin_notifs.first().message)
        self.assertEqual(admin_notifs.first().url, reverse('dashboard_admin') + '?tab=students')

        # Check notifications generated for superuser
        super_notifs = Notification.objects.filter(destinataire=self.superuser)
        self.assertEqual(super_notifs.count(), 1)
        self.assertEqual(super_notifs.first().type, 'NOUVELLE_INSCRIPTION')

    def test_student_activation_deletes_registration_notification(self):
        """Test that activating a pending student deletes their registration notifications."""
        from accounts.models import Niveau, Classe

        # 1. Register student
        self.client.post(reverse('accounts:register'), {
            'email': 'new_student_test@example.com',
            'password1': 'NewStudentPass123!',
            'password2': 'NewStudentPass123!',
            'role': 'ELEVE'
        })

        # Check notification exists initially
        admin_notifs = Notification.objects.filter(destinataire=self.admin, type='NOUVELLE_INSCRIPTION')
        self.assertEqual(admin_notifs.count(), 1)
        self.assertFalse(admin_notifs.first().lu)

        # 2. Setup Niveau and Classe
        niveau = Niveau.objects.create(code="nv1", nom="Niveau 1", ordre=1)
        classe = Classe.objects.create(niveau=niveau, code="cls1", nom="Classe 1", annee_scolaire="2026", capacite=30)

        # 3. Log in as admin
        self.client.login(email="admin_test_notif@example.com", password="AdminPassword123!")

        # 4. Get the student object
        student = Utilisateur.objects.get(email='new_student_test@example.com')

        # 5. Activate the student via the POST view
        response = self.client.post(reverse('dashboard_admin_activate_pending_student'), {
            'student_id': str(student.id),
            'classe': str(classe.id)
        })
        self.assertEqual(response.status_code, 200)

        # 6. Verify notification is deleted
        admin_notifs = Notification.objects.filter(destinataire=self.admin, type='NOUVELLE_INSCRIPTION')
        self.assertEqual(admin_notifs.count(), 0)


class NotificationHTMXTests(TestCase):
    def setUp(self):
        self.user = Utilisateur.objects.create_user(
            email="notif_test@example.com",
            password="Password123!",
            role="ELEVE",
            statut_compte="ACTIVE",
            is_active=True
        )
        self.notif1 = Notification.objects.create(
            destinataire=self.user,
            type="COMPTE_ACTIVE",
            titre="Test Notif 1",
            message="Msg 1"
        )
        self.notif2 = Notification.objects.create(
            destinataire=self.user,
            type="COMPTE_ACTIVE",
            titre="Test Notif 2",
            message="Msg 2"
        )

    def test_unread_notifications_count_view(self):
        self.client.login(email="notif_test@example.com", password="Password123!")
        
        # Initial check (2 unread)
        response = self.client.get(reverse('accounts:unread_notifications_count'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'notifications-badge')
        self.assertContains(response, '2')

        # Mark one as read
        self.notif1.lu = True
        self.notif1.save()

        # Check count is updated to 1
        response = self.client.get(reverse('accounts:unread_notifications_count'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '1')

        # Mark second as read
        self.notif2.lu = True
        self.notif2.save()

        # Check count is 0 (returns empty response, no badge)
        response = self.client.get(reverse('accounts:unread_notifications_count'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode(), '')

    def test_mark_notification_read_view_htmx(self):
        self.client.login(email="notif_test@example.com", password="Password123!")
        
        # Test marking one notification read via HTMX
        response = self.client.post(
            reverse('accounts:mark_notification_read', args=[self.notif1.id]),
            headers={'hx-request': 'true'}
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['HX-Trigger'], 'update-notifications')
        
        self.notif1.refresh_from_db()
        self.assertTrue(self.notif1.lu)

    def test_mark_notifications_read_all_view_htmx(self):
        self.client.login(email="notif_test@example.com", password="Password123!")
        
        # Test marking all notifications read via HTMX
        response = self.client.post(
            reverse('accounts:mark_notifications_read_all'),
            headers={'hx-request': 'true'}
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['HX-Trigger'], 'update-notifications')
        
        self.notif1.refresh_from_db()
        self.notif2.refresh_from_db()
        self.assertTrue(self.notif1.lu)
        self.assertTrue(self.notif2.lu)

    def test_read_and_redirect_notification_view(self):
        self.client.login(email="notif_test@example.com", password="Password123!")
        self.notif1.url = "/some-target-url/"
        self.notif1.lu = False
        self.notif1.save()

        response = self.client.get(reverse('accounts:read_and_redirect_notification', args=[self.notif1.id]))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, "/some-target-url/")

        self.notif1.refresh_from_db()
        self.assertTrue(self.notif1.lu)

    def test_delete_notification_view(self):
        self.client.login(email="notif_test@example.com", password="Password123!")
        notif_id = self.notif1.id
        
        response = self.client.post(
            reverse('accounts:delete_notification', args=[notif_id]),
            headers={'hx-request': 'true'}
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['HX-Trigger'], 'update-notifications')
        
        # Verify notification is deleted
        self.assertFalse(Notification.objects.filter(id=notif_id).exists())


from accounts.models import Niveau, Classe

class ClassStudentsManagementTests(TestCase):
    def setUp(self):
        # Setup Admin
        self.admin = Utilisateur.objects.create_user(
            email="admin_class_test@example.com",
            password="AdminPassword123!",
            role="ADMIN",
            statut_compte="ACTIVE",
            is_active=True
        )
        # Setup Niveau and Classe
        self.niveau = Niveau.objects.create(code="nv_test", nom="Niveau Test", ordre=1)
        self.classe = Classe.objects.create(niveau=self.niveau, code="cls_test", nom="Classe Test", annee_scolaire="2025-2026", capacite=30)
        
        # Setup Student in Classe
        self.student_in_class = Utilisateur.objects.create_user(
            email="in_class@example.com",
            password="StudentPass123!",
            role="ELEVE",
            statut_compte="ACTIVE",
            is_active=True,
            classe=self.classe
        )
        
        # Setup Pending Student (no class)
        self.pending_student = Utilisateur.objects.create_user(
            email="pending_stu@example.com",
            password="StudentPass123!",
            role="ELEVE",
            statut_compte="PENDING",
            is_active=False
        )

    def test_manage_classe_students_view_unauthenticated(self):
        """Unauthenticated user is redirected to login."""
        response = self.client.get(reverse('dashboard_admin_manage_classe_students', args=[self.classe.id]))
        self.assertEqual(response.status_code, 302)

    def test_manage_classe_students_view_forbidden_for_student(self):
        """A student cannot access class management."""
        self.client.login(email="in_class@example.com", password="StudentPass123!")
        response = self.client.get(reverse('dashboard_admin_manage_classe_students', args=[self.classe.id]))
        self.assertEqual(response.status_code, 403)

    def test_manage_classe_students_view_ok_for_admin(self):
        """Admin can access the class students page."""
        self.client.login(email="admin_class_test@example.com", password="AdminPassword123!")
        response = self.client.get(reverse('dashboard_admin_manage_classe_students', args=[self.classe.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.classe.nom)
        self.assertContains(response, "in_class@example.com")
        self.assertContains(response, "pending_stu@example.com")

    def test_export_classe_students_excel(self):
        """Admin can export class students list to excel."""
        self.client.login(email="admin_class_test@example.com", password="AdminPassword123!")
        response = self.client.get(reverse('dashboard_admin_export_classe_students', args=[self.classe.id]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn(f'attachment; filename="eleves_{self.classe.nom}.xlsx"', response['Content-Disposition'])
        
        # Verify it is a valid openpyxl workbook
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIn("Élèves", wb.sheetnames)
        ws = wb["Élèves"]
        # Check title in A1
        self.assertEqual(ws['A1'].value, f"Liste des élèves - Classe : {self.classe.nom}")









