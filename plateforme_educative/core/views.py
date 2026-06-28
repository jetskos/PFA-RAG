import json
from datetime import timedelta
from django.utils import timezone
from functools import wraps
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import AccessMixin, LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponseForbidden, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.views.decorators.http import require_http_methods

from apprentissage.models import Cours
from accounts.models import Classe, Niveau
from logistics.forms import DemandeMaterielForm
from logistics.models import DemandeMateriel, Equipment
from .forms import NiveauForm, ClasseForm, PendingStudentActivationForm

Utilisateur = get_user_model()


def home_view(request):
    """Affiche la Landing Page publique. Redirige les utilisateurs connectés vers leur dashboard."""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    return render(request, 'core/home.html')


def _admin_dashboard_context(request=None):
    from django.db.models import Q
    pending_students = Utilisateur.objects.filter(role='ELEVE', statut_compte='PENDING').order_by('-date_creation')
    pending_rows = [
        {
            'user': student,
            'form': PendingStudentActivationForm(initial={'student_id': student.id}),
        }
        for student in pending_students
    ]
    
    # Calculate stats for the dashboard counters
    eleves_actifs_count = Utilisateur.objects.filter(role='ELEVE', is_active=True).count()
    formateurs_actifs_count = Utilisateur.objects.filter(role='FORMATEUR', is_active=True).count()
    en_attente_count = pending_students.count()
    cours_publies_count = Cours.objects.filter(actif=True).count()
    tickets_gmao_count = DemandeMateriel.objects.filter(statut='PENDING').count()
    niveaux_actifs_count = Niveau.objects.filter(actif=True).count()

    # Generate a unified timeline from recent objects
    timeline = []
    
    # 1. Recent tickets
    recent_tickets = DemandeMateriel.objects.select_related('formateur', 'equipement').order_by('-date_creation')[:3]
    for t in recent_tickets:
        timeline.append({
            'date': t.date_creation,
            'type': 'ticket',
            'dot_class': 'dot-red',
            'title': 'Nouveau ticket matériel ouvert',
            'desc': f"Le formateur {t.formateur.get_full_name()} a signalé un besoin pour {t.equipement.nom if t.equipement else 'du matériel'}."
        })
        
    # 2. Recent courses
    recent_courses = Cours.objects.select_related('createur', 'niveau').order_by('-date_creation')[:3]
    for c in recent_courses:
        timeline.append({
            'date': c.date_creation,
            'type': 'cours',
            'dot_class': 'dot-blue',
            'title': 'Cours publié',
            'desc': f"Le module '{c.titre}' a été mis en ligne pour le niveau {c.niveau.nom}."
        })
        
    # 3. Recent users
    recent_users = Utilisateur.objects.order_by('-date_creation')[:3]
    for u in recent_users:
        timeline.append({
            'date': u.date_creation,
            'type': 'user',
            'dot_class': 'dot-green',
            'title': 'Nouvel utilisateur inscrit',
            'desc': f"L'utilisateur {u.get_full_name()} a créé son compte."
        })
        
    # Sort timeline by date descending and limit to 5
    timeline.sort(key=lambda x: x['date'], reverse=True)
    timeline = timeline[:5]

    # Filtering logic for lists
    niveaux_qs = Niveau.objects.all().order_by('ordre')
    classes_qs = Classe.objects.select_related('niveau').all().order_by('nom')
    pending_students_qs = Utilisateur.objects.filter(role='ELEVE', statut_compte='PENDING').order_by('-date_creation')
    pending_demandes_qs = DemandeMateriel.objects.filter(statut='PENDING').select_related('formateur', 'equipement', 'atelier_cible').order_by('-date_creation')
    equipements_qs = Equipment.objects.all().order_by('nom')
    
    if request:
        q = request.GET.get('q', '').strip()
        status_filter = request.GET.get('status', '')
        
        if q:
            niveaux_qs = niveaux_qs.filter(Q(nom__icontains=q) | Q(code__icontains=q))
            classes_qs = classes_qs.filter(Q(nom__icontains=q) | Q(code__icontains=q) | Q(niveau__nom__icontains=q))
            pending_students_qs = pending_students_qs.filter(Q(email__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q))
            pending_demandes_qs = pending_demandes_qs.filter(Q(description__icontains=q) | Q(formateur__email__icontains=q) | Q(equipement__nom__icontains=q))
            
        if status_filter == 'actif':
            niveaux_qs = niveaux_qs.filter(actif=True)
            classes_qs = classes_qs.filter(actif=True)
        elif status_filter == 'inactif':
            niveaux_qs = niveaux_qs.filter(actif=False)
            classes_qs = classes_qs.filter(actif=False)
            
        page_number = request.GET.get('page', 1)
    else:
        page_number = 1

    # Pagination
    from django.core.paginator import Paginator
    
    niveaux_paginator = Paginator(niveaux_qs, 10)
    niveaux_page = niveaux_paginator.get_page(page_number)
    
    classes_paginator = Paginator(classes_qs, 10)
    classes_page = classes_paginator.get_page(page_number)
    
    students_paginator = Paginator(pending_students_qs, 10)
    pending_students_page = students_paginator.get_page(page_number)
    
    demandes_paginator = Paginator(pending_demandes_qs, 10)
    pending_demandes_page = demandes_paginator.get_page(page_number)
    
    equipements_paginator = Paginator(equipements_qs, 10)
    equipements_page = equipements_paginator.get_page(page_number)

    pending_rows = [
        {
            'user': student,
            'form': PendingStudentActivationForm(initial={'student_id': student.id}),
        }
        for student in pending_students_page
    ]

    # === CHART DATA GENERATION ===
    admins_count = Utilisateur.objects.filter(role='ADMIN').count()
    users_chart = {
        'students': eleves_actifs_count + en_attente_count,
        'formateurs': formateurs_actifs_count,
        'admins': admins_count
    }

    tickets_ouvert = DemandeMateriel.objects.filter(statut='PENDING').count()
    tickets_encours = DemandeMateriel.objects.filter(statut='APPROVED').count()  # Adjust if EN_COURS exists
    tickets_resolu = DemandeMateriel.objects.filter(statut='REJECTED').count()   # Adjust if RESOLU exists
    tickets_chart = {
        'ouvert': tickets_ouvert,
        'encours': tickets_encours,
        'resolu': tickets_resolu
    }

    today = timezone.now().date()
    labels = []
    values = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        count = Utilisateur.objects.filter(date_creation__date=day).count()
        labels.append(day.strftime('%d/%m'))
        values.append(count)
    inscriptions_chart = {
        'labels': labels,
        'values': values
    }

    chart_data = {
        'users': users_chart,
        'tickets': tickets_chart,
        'inscriptions': inscriptions_chart
    }

    return {
        'niveaux': niveaux_page,
        'classes': classes_page,
        'equipements': equipements_page,
        'pending_students': pending_students_page,
        'pending_rows': pending_rows,
        'pending_demandes': pending_demandes_page,
        'niveau_form': NiveauForm(),
        'classe_form': ClasseForm(),
        'niveau_editor_form': None,
        'niveau_editor_title': '',
        'stats': {
            'eleves_actifs': eleves_actifs_count,
            'formateurs_actifs': formateurs_actifs_count,
            'en_attente': en_attente_count,
            'cours_publies': cours_publies_count,
            'tickets_gmao': tickets_gmao_count,
            'niveaux_actifs': niveaux_actifs_count,
        },
        'timeline': timeline,
        'chart_data': chart_data,
    }


def _render_admin_structure(request, context=None, status=200, active_tab='niveaux'):
    render_context = _admin_dashboard_context(request)
    render_context['active_tab'] = active_tab
    if context:
        render_context.update(context)
    return render(request, 'core/partials/admin_structure.html', render_context, status=status)


def _formateur_dashboard_context(request):
    mes_cours = Cours.objects.filter(createur=request.user).select_related('niveau').order_by('-date_creation')
    mes_demandes = DemandeMateriel.objects.filter(formateur=request.user).select_related(
        'equipement',
        'atelier_cible',
    ).order_by('-date_creation')
    
    from logistics.models import Workshop
    from apprentissage.models import Chapitre
    
    mes_ateliers = Workshop.objects.filter(createur=request.user).order_by('-date_debut')
    
    total_cours = mes_cours.count()
    total_chapitres = Chapitre.objects.filter(cours__createur=request.user).count()
    demandes_en_attente = DemandeMateriel.objects.filter(formateur=request.user, statut='PENDING').count()
    
    return {
        'mes_cours': mes_cours,
        'mes_demandes': mes_demandes,
        'mes_ateliers': mes_ateliers,
        'total_cours': total_cours,
        'total_chapitres': total_chapitres,
        'demandes_en_attente': demandes_en_attente,
        'demande_form': DemandeMaterielForm(),
    }


def _render_formateur_structure(request, context=None, status=200):
    render_context = _formateur_dashboard_context(request)
    if context:
        render_context.update(context)
    return render(request, 'core/partials/formateur_structure.html', render_context, status=status)


def role_required(*allowed_roles):
    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def _wrapped(request, *args, **kwargs):
            user = request.user
            if not getattr(user, 'is_active', False):
                return HttpResponseForbidden('Votre compte est en attente de validation.')
            if user.role not in allowed_roles and not user.is_superuser:
                return HttpResponseForbidden('Accès refusé.')
            return view_func(request, *args, **kwargs)

        return _wrapped

    return decorator


class RoleRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    allowed_roles = ()

    def test_func(self):
        user = self.request.user
        if not getattr(user, 'is_active', False):
            return False
        return user.is_superuser or user.role in self.allowed_roles


def dashboard_router(request):
    if not request.user.is_authenticated:
        return redirect('accounts:login')

    if not getattr(request.user, 'is_active', False):
        return HttpResponseForbidden('Votre compte est en attente de validation.')

    if request.user.is_superuser or request.user.role == 'ADMIN':
        return redirect('dashboard_admin')
    if request.user.role == 'FORMATEUR':
        return redirect('dashboard_formateur')
    return redirect('dashboard_student')


@role_required('ADMIN')
def admin_dashboard_view(request):
    context = {
        'role': 'ADMIN',
        'active_tab': 'command_center',
        **_admin_dashboard_context(request),
    }
    # La recherche HTMX sur le dashboard sera gérée spécifiquement plus tard
    return render(request, 'core/dashboard_admin.html', context)


@role_required('ADMIN')
def admin_structure_page_view(request):
    tab = request.GET.get('tab', 'niveaux')
    context = {
        'role': 'ADMIN',
        'active_tab': tab,
        **_admin_dashboard_context(request),
    }
    if request.headers.get('HX-Request') and request.headers.get('HX-Target') != 'zone-principale':
        return render(request, 'core/partials/admin_structure.html', context)
    return render(request, 'core/dashboard_admin_structure.html', context)


@role_required('FORMATEUR')
def formateur_dashboard_view(request):
    tab = request.GET.get('tab', 'all')
    context = {
        'role': 'FORMATEUR',
        'active_tab': tab,
        **_formateur_dashboard_context(request),
    }
    if request.headers.get('HX-Request') and request.headers.get('HX-Target') != 'zone-principale':
        return render(request, 'core/partials/formateur_structure.html', context)
    return render(request, 'core/dashboard_formateur.html', context)


@role_required('ELEVE')
def student_dashboard_view(request):
    import json
    from apprentissage.models import Progression

    classe = getattr(request.user, 'classe', None)
    niveau = getattr(classe, 'niveau', None) if classe else None
    
    cours_data = []
    radar_labels = []
    radar_data = []

    if niveau:
        mes_cours = Cours.objects.filter(niveau=niveau, actif=True).order_by('titre')
        for cours in mes_cours:
            progression_obj = Progression.objects.filter(etudiant=request.user, cours=cours).first()
            progression_percent = progression_obj.pourcentage if progression_obj else 0
            
            cours_data.append({
                'id': cours.id,
                'titre': cours.titre,
                'description': cours.description,
                'chapitres_count': cours.chapitres.count(),
                'progression_percent': progression_percent,
                'image_couverture': cours.image_couverture
            })
            
            radar_labels.append(cours.titre[:15] + '...' if len(cours.titre) > 15 else cours.titre)
            radar_data.append(progression_percent)

    if not radar_labels:
        radar_labels = ['Aucun cours']
        radar_data = [0]

    return render(request, 'core/dashboard_student.html', {
        'role': 'ELEVE',
        'classe': classe,
        'niveau': niveau,
        'mes_cours': cours_data,
        'radar_labels_json': json.dumps(radar_labels),
        'radar_data_json': json.dumps(radar_data),
    })


@role_required('ADMIN')
@require_http_methods(['GET', 'POST'])
def create_niveau_view(request):
    if request.method == 'POST':
        form = NiveauForm(request.POST)
        if form.is_valid():
            form.save()
            response = _render_admin_structure(request, active_tab='niveaux')
            response['HX-Trigger'] = 'closeModal'
            response['HX-Retarget'] = '#dashboard-structure'
            return response
    else:
        form = NiveauForm()

    if request.headers.get('HX-Request'):
        return render(request, 'core/partials/niveau_form.html', {
            'niveau_form': form,
            'submit_label': 'Créer le niveau',
            'cancel_label': 'Annuler',
            'action_url': reverse('dashboard_admin_create_niveau'),
            'is_create': True
        }, status=400 if request.method == 'POST' else 200)
    
    return _render_admin_structure(request, {'niveau_form': form}, status=400 if request.method == 'POST' else 200, active_tab='niveaux')


@role_required('ADMIN')
@require_http_methods(['GET', 'POST'])
def edit_niveau_view(request, niveau_id):
    niveau = get_object_or_404(Niveau, pk=niveau_id)

    if request.method == 'POST':
        form = NiveauForm(request.POST, instance=niveau)
        if form.is_valid():
            form.save()
            response = _render_admin_structure(request, active_tab='niveaux')
            response['HX-Trigger'] = 'closeModal'
            response['HX-Retarget'] = '#dashboard-structure'
            return response
    else:
        form = NiveauForm(instance=niveau)

    return render(request, 'core/partials/niveau_form.html', {
        'niveau_form': form,
        'niveau': niveau,
        'submit_label': 'Enregistrer',
        'cancel_label': 'Annuler',
        'action_url': reverse('dashboard_admin_edit_niveau', args=[niveau.id]),
    }, status=400 if request.method == 'POST' else 200)


@role_required('ADMIN')
@require_http_methods(['GET', 'POST'])
def create_classe_view(request):
    if request.method == 'POST':
        form = ClasseForm(request.POST)
        if form.is_valid():
            form.save()
            response = _render_admin_structure(request, active_tab='classes')
            response['HX-Trigger'] = 'closeModal'
            response['HX-Retarget'] = '#dashboard-structure'
            return response
    else:
        form = ClasseForm()

    if request.headers.get('HX-Request'):
        return render(request, 'core/partials/classe_form.html', {
            'classe_form': form,
            'submit_label': 'Créer la classe',
            'cancel_label': 'Annuler',
            'action_url': reverse('dashboard_admin_create_classe'),
            'is_create': True
        }, status=400 if request.method == 'POST' else 200)

    return _render_admin_structure(request, {'classe_form': form}, status=400 if request.method == 'POST' else 200, active_tab='classes')


@role_required('ADMIN')
@require_http_methods(['GET', 'POST'])
def edit_classe_view(request, classe_id):
    classe = get_object_or_404(Classe, pk=classe_id)

    if request.method == 'POST':
        form = ClasseForm(request.POST, instance=classe)
        if form.is_valid():
            form.save()
            response = _render_admin_structure(request, active_tab='classes')
            response['HX-Trigger'] = 'closeModal'
            response['HX-Retarget'] = '#dashboard-structure'
            return response
    else:
        form = ClasseForm(instance=classe)

    return render(request, 'core/partials/classe_form.html', {
        'classe_form': form,
        'classe': classe,
        'submit_label': 'Enregistrer',
        'cancel_label': 'Annuler',
        'action_url': reverse('dashboard_admin_edit_classe', args=[classe.id]),
    }, status=400 if request.method == 'POST' else 200)


@role_required('ADMIN')
@require_http_methods(['GET'])
def manage_classe_students_view(request, classe_id):
    classe = get_object_or_404(Classe, pk=classe_id)
    eleves = Utilisateur.objects.filter(classe=classe, role='ELEVE').order_by('email')
    pending_students = Utilisateur.objects.filter(classe__isnull=True, role='ELEVE').order_by('email')
    return render(request, 'core/classe_students_page.html', {
        'classe': classe,
        'eleves': eleves,
        'pending_students': pending_students,
    })

@role_required('ADMIN')
@require_http_methods(['GET'])
def export_classe_students_excel_view(request, classe_id):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    classe = get_object_or_404(Classe, pk=classe_id)
    eleves = Utilisateur.objects.filter(classe=classe, role='ELEVE').order_by('last_name', 'first_name', 'email')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Élèves"

    # Show gridlines explicitly
    ws.views.sheetView[0].showGridLines = True

    # Title block
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Liste des élèves - Classe : {classe.nom}"
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid") # Teal brand
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Metadata
    ws['A2'] = "Niveau :"
    ws['B2'] = classe.niveau.nom
    ws['A3'] = "Année Scolaire :"
    ws['B3'] = classe.annee_scolaire
    ws['A4'] = "Effectif :"
    ws['B4'] = f"{eleves.count()} élève(s)"

    meta_label_font = Font(name="Calibri", size=11, bold=True, color="334155")
    meta_val_font = Font(name="Calibri", size=11, color="334155")
    for row in range(2, 5):
        ws.cell(row=row, column=1).font = meta_label_font
        ws.cell(row=row, column=2).font = meta_val_font
    
    ws.row_dimensions[5].height = 15

    # Headers
    headers = ["Nom", "Prénom", "Email", "Statut Compte", "Date d'Inscription"]
    header_row = 6
    ws.row_dimensions[header_row].height = 25
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo secondary
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data
    row_idx = 7
    data_font = Font(name="Calibri", size=11)
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    for student in eleves:
        ws.row_dimensions[row_idx].height = 20
        status = student.get_statut_compte_display() if hasattr(student, 'get_statut_compte_display') else ("Actif" if student.is_active else "Inactif")
        date_str = student.date_creation.strftime("%d/%m/%Y %H:%M") if student.date_creation else ""

        row_data = [
            student.last_name or "-",
            student.first_name or "-",
            student.email,
            status,
            date_str
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in [3, 4, 5]:
                cell.alignment = align_center if col_idx != 3 else align_left
            else:
                cell.alignment = align_left
            
            # Zebra
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        row_idx += 1

    # Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 1 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="eleves_{classe.nom}.xlsx"'
    wb.save(response)
    return response

@role_required('ADMIN')
@require_http_methods(['POST'])
def remove_student_from_classe_page_view(request, classe_id, student_id):
    student = get_object_or_404(Utilisateur, pk=student_id)
    student.classe = None
    student.save(update_fields=['classe'])
    return redirect('dashboard_admin_manage_classe_students', classe_id=classe_id)

@role_required('ADMIN')
@require_http_methods(['POST'])
def delete_student_from_classe_page_view(request, classe_id, student_id):
    student = get_object_or_404(Utilisateur, pk=student_id)
    # Clear registration notifications for this student before deleting
    from accounts.models import Notification
    Notification.objects.filter(
        type='NOUVELLE_INSCRIPTION',
        message__contains=student.email
    ).delete()
    student.delete()
    return redirect('dashboard_admin_manage_classe_students', classe_id=classe_id)

@role_required('ADMIN')
@require_http_methods(['POST'])
def assign_student_to_classe_page_view(request, classe_id):
    classe = get_object_or_404(Classe, pk=classe_id)
    student_id = request.POST.get('student_id')
    if student_id:
        student = get_object_or_404(Utilisateur, pk=student_id)
        student.classe = classe
        student.statut_compte = 'ACTIVE'
        student.is_active = True
        student.save(update_fields=['classe', 'statut_compte', 'is_active'])
        
        # Clear registration notifications for this student
        from accounts.models import Notification
        Notification.objects.filter(
            type='NOUVELLE_INSCRIPTION',
            message__contains=student.email
        ).delete()

        from accounts.notifications import notifier
        notifier(
            destinataire=student,
            type='COMPTE_ACTIVE',
            titre="Votre compte a été activé",
            message=f"Votre inscription a été validée. Vous êtes maintenant affecté à la classe {classe.nom}.",
            envoyer_email=True
        )
    return redirect('dashboard_admin_manage_classe_students', classe_id=classe_id)

@role_required('ADMIN')
@require_http_methods(['POST'])
def add_student_to_classe_page_view(request, classe_id):
    classe = get_object_or_404(Classe, pk=classe_id)
    email = request.POST.get('email')
    password = request.POST.get('password')
    first_name = request.POST.get('first_name', '').strip()
    last_name = request.POST.get('last_name', '').strip()
    if email and password and not Utilisateur.objects.filter(email=email).exists():
        Utilisateur.objects.create_user(
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            role='ELEVE',
            statut_compte='ACTIVE',
            is_active=True,
            classe=classe
        )
    return redirect('dashboard_admin_manage_classe_students', classe_id=classe_id)


@role_required('ADMIN')
@require_http_methods(['GET', 'POST'])
def activate_pending_student_view(request):
    if request.method != 'POST':
        return HttpResponseForbidden('Méthode non autorisée.')

    form = PendingStudentActivationForm(request.POST)
    if form.is_valid():
        student = Utilisateur.objects.get(pk=form.cleaned_data['student_id'])
        student.classe = form.cleaned_data['classe']
        student.statut_compte = 'ACTIVE'
        student.is_active = True
        if student.role == 'ELEVE':
            student.is_formateur = False
        student.save(update_fields=['classe', 'statut_compte', 'is_active', 'is_formateur'])
        
        # Clear registration notifications for this student
        from accounts.models import Notification
        Notification.objects.filter(
            type='NOUVELLE_INSCRIPTION',
            message__contains=student.email
        ).delete()

        from accounts.notifications import notifier
        notifier(
            destinataire=student,
            type='COMPTE_ACTIVE',
            titre="Votre compte a été activé",
            message=f"Félicitations ! Votre compte est désormais actif. Vous êtes affecté à la classe {student.classe.nom}.",
            envoyer_email=True
        )
        return _render_admin_structure(request, active_tab='students')

    return _render_admin_structure(request, {
        'pending_error': form.errors.as_text(),
    }, status=400, active_tab='students')


@role_required('FORMATEUR')
@require_http_methods(['GET', 'POST'])
def create_demande_materiel_view(request):
    if request.method == 'POST':
        form = DemandeMaterielForm(request.POST)
        if form.is_valid():
            demande = form.save(commit=False)
            demande.formateur = request.user
            demande.save()
            # On success, close modal and show toast (empty modal container + trigger JS event if needed)
            response_html = """
<div id="modal-container"></div>
<div id="htmx-toast" style="
    position: fixed;
    bottom: 2rem;
    right: 2rem;
    z-index: 9999;
    background: #10b981;
    color: white;
    padding: 1rem 1.5rem;
    border-radius: 12px;
    box-shadow: 0 10px 25px -5px rgba(16,185,129,0.4);
    display: flex;
    align-items: center;
    gap: 0.75rem;
    font-weight: 600;
    font-size: 0.95rem;
    animation: slideInRight 0.3s cubic-bezier(0.16,1,0.3,1);
    max-width: 380px;">
    <span style="font-size:1.25rem;">✅</span>
    <div>
        <div style="font-weight:700;">Demande envoyée !</div>
        <div style="font-size:0.85rem;opacity:0.9;font-weight:400;">L'administrateur recevra votre demande de matériel.</div>
    </div>
</div>
<style>
@keyframes slideInRight {
    from { opacity: 0; transform: translateX(100px); }
    to { opacity: 1; transform: translateX(0); }
}
</style>
<script>
    setTimeout(() => {
        const toast = document.getElementById('htmx-toast');
        if (toast) {
            toast.style.transition = 'opacity 0.4s ease, transform 0.4s ease';
            toast.style.opacity = '0';
            toast.style.transform = 'translateX(100px)';
            setTimeout(() => toast.remove(), 400);
        }
    }, 4000);
</script>
"""
            return HttpResponse(response_html)
    else:
        form = DemandeMaterielForm()

    # If GET or invalid POST, render the modal
    context = {'demande_form': form}
    return render(request, 'core/partials/demande_materiel_modal.html', context, status=400 if request.method == 'POST' else 200)


from django.db import transaction

@role_required('ADMIN')
@require_http_methods(['POST'])
@transaction.atomic
def admin_process_demande_view(request, demande_id):
    """Process a DemandeMateriel: approve, reject or return. Returns refreshed pending/active demandes section."""
    action = request.POST.get('action')
    demande = get_object_or_404(DemandeMateriel, pk=demande_id)

    try:
        with transaction.atomic():
            equipement = demande.equipement
            
            if action == 'approve':
                if demande.quantite > equipement.stock_disponible:
                    return _render_admin_structure(request, {'pending_error': f'Stock insuffisant (Reste: {equipement.stock_disponible}).'}, status=400, active_tab='logistics')
                equipement.stock_disponible -= demande.quantite
                equipement.save()
                demande.statut = 'APPROVED'
            elif action == 'reject':
                demande.statut = 'REJECTED'
            elif action == 'return':
                if demande.statut == 'APPROVED':
                    equipement.stock_disponible += demande.quantite
                    equipement.save()
                demande.statut = 'RETURNED'
            else:
                return _render_admin_structure(request, {'pending_error': 'Action invalide.'}, status=400, active_tab='logistics')

            demande.save(update_fields=['statut'])
    except Exception as e:
        return _render_admin_structure(request, {'pending_error': str(e)}, status=400, active_tab='logistics')

    from accounts.notifications import notifier
    if action in ['approve', 'reject']:
        statut_str = "approuvée" if action == 'approve' else "rejetée"
        notifier(
            destinataire=demande.formateur,
            type='DEMANDE_TRAITEE',
            titre=f"Demande de matériel {statut_str}",
            message=f"Votre demande pour l'équipement '{demande.equipement.nom}' a été {statut_str} par l'administrateur.",
            envoyer_email=True
        )

    # After processing, re-render the appropriate sections so HTMX can swap them.
    # Currently it just swapped the pending_demandes_section
    context = {
        'pending_demandes': DemandeMateriel.objects.filter(statut='PENDING').select_related('formateur', 'equipement', 'atelier_cible').order_by('-date_creation'),
        'active_demandes': DemandeMateriel.objects.filter(statut='APPROVED').select_related('formateur', 'equipement', 'atelier_cible').order_by('-date_mise_a_jour'),
    }
    return render(request, 'core/partials/pending_demandes_section.html', context)


@role_required('ADMIN')
@require_http_methods(['GET', 'POST'])
def admin_equipement_create_view(request):
    from logistics.forms import EquipmentForm
    if request.method == 'POST':
        form = EquipmentForm(request.POST)
        if form.is_valid():
            form.save()
            response = _render_admin_structure(request, active_tab='equipements')
            response['HX-Trigger'] = 'closeModal'
            response['HX-Retarget'] = '#dashboard-structure'
            return response
    else:
        form = EquipmentForm()
    return render(request, 'core/partials/equipement_form_modal.html', {'equipement_form': form, 'is_update': False})


@role_required('ADMIN')
@require_http_methods(['GET', 'POST'])
def admin_equipement_update_view(request, equipement_id):
    from logistics.models import Equipment
    from logistics.forms import EquipmentForm
    equipement = get_object_or_404(Equipment, pk=equipement_id)
    if request.method == 'POST':
        form = EquipmentForm(request.POST, instance=equipement)
        if form.is_valid():
            form.save()
            return render(request, 'core/partials/equipement_row.html', {'equipement': equipement})
    else:
        form = EquipmentForm(instance=equipement)
    return render(request, 'core/partials/equipement_form_modal.html', {'equipement_form': form, 'is_update': True, 'equipement': equipement})


@role_required('ADMIN')
@require_http_methods(['POST'])
def admin_equipement_toggle_active_view(request, equipement_id):
    from logistics.models import Equipment
    equipement = get_object_or_404(Equipment, pk=equipement_id)
    equipement.est_actif = not equipement.est_actif
    equipement.save()
    return render(request, 'core/partials/equipement_row.html', {'equipement': equipement})