"""
Utilitaires pour exporter les données Logistics au format Excel formaté.
"""
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# Couleurs personnalisées
COLOR_HEADER = "0F766E"  # Teal
COLOR_ACTIVE = "10B981"  # Emerald (Actif)
COLOR_INACTIVE = "EF4444"  # Red (Inactif)
COLOR_PENDING = "F59E0B"  # Amber (En attente)
COLOR_APPROVED = "10B981"  # Emerald (Approuvé)
COLOR_REJECTED = "EF4444"  # Red (Rejeté)


class ExcelExporter:
    """Classe pour générer des fichiers Excel formatés."""

    def __init__(self, title, filename_suffix):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Données"
        self.title = title
        self.filename_suffix = filename_suffix
        self.current_row = 1
        self.headers = []
        self.rows = []

    def add_title(self, title_text):
        """Ajouter un titre au haut du document."""
        self.worksheet.merge_cells("A1:H1")
        title_cell = self.worksheet["A1"]
        title_cell.value = title_text
        title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        self.worksheet.row_dimensions[1].height = 30
        self.current_row = 2
        return self

    def add_subtitle(self, subtitle_text):
        """Ajouter un sous-titre."""
        self.worksheet.merge_cells(f"A{self.current_row}:H{self.current_row}")
        subtitle_cell = self.worksheet[f"A{self.current_row}"]
        subtitle_cell.value = f"📅 Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        subtitle_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
        self.current_row += 1
        return self

    def add_headers(self, headers):
        """Ajouter les en-têtes du tableau."""
        self.headers = headers
        self.current_row += 1  # Ligne vide
        
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_num)
            cell.value = header
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )
        
        self.worksheet.row_dimensions[self.current_row].height = 25
        header_row = self.current_row
        self.current_row += 1
        
        # Figeage des en-têtes
        self.worksheet.freeze_panes = f"A{self.current_row}"
        
        return header_row

    def add_row(self, row_data, status_col=None, status_value=None):
        """Ajouter une ligne de données."""
        for col_num, value in enumerate(row_data, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_num)
            cell.value = value
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
                top=Side(style="thin", color="E2E8F0"),
                bottom=Side(style="thin", color="E2E8F0"),
            )
            
            # Formatage des statuts
            if status_col is not None and col_num == status_col:
                cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if status_value == "Actif":
                    cell.fill = PatternFill(start_color=COLOR_ACTIVE, end_color=COLOR_ACTIVE, fill_type="solid")
                elif status_value in ["Inactif", "Annulé"]:
                    cell.fill = PatternFill(start_color=COLOR_INACTIVE, end_color=COLOR_INACTIVE, fill_type="solid")
                elif status_value == "En attente":
                    cell.fill = PatternFill(start_color=COLOR_PENDING, end_color=COLOR_PENDING, fill_type="solid")
                elif status_value == "Approuvée":
                    cell.fill = PatternFill(start_color=COLOR_APPROVED, end_color=COLOR_APPROVED, fill_type="solid")
                elif status_value == "Rejetée":
                    cell.fill = PatternFill(start_color=COLOR_REJECTED, end_color=COLOR_REJECTED, fill_type="solid")
        
        self.worksheet.row_dimensions[self.current_row].height = 20
        self.current_row += 1
        return self

    def auto_resize_columns(self):
        """Redimensionner automatiquement les colonnes."""
        for col_num, header in enumerate(self.headers, 1):
            max_length = len(str(header)) + 2
            
            for row in self.worksheet.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)) + 2)
            
            adjusted_width = min(max_length, 50)  # Max 50 chars
            column_letter = get_column_letter(col_num)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return self

    def add_autofilter(self, header_row):
        """Ajouter les filtres automatiques."""
        if self.headers:
            last_col = get_column_letter(len(self.headers))
            self.worksheet.auto_filter.ref = f"A{header_row}:{last_col}{header_row}"
        return self

    def get_bytes(self):
        """Retourner le fichier en bytes."""
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def get_filename(self):
        """Retourner le nom du fichier."""
        return f"{self.filename_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def export_equipements_excel(equipements):
    """Exporter les équipements en Excel formaté."""
    exporter = ExcelExporter("Équipements", "equipements")
    exporter.add_title("📦 Gestion des Équipements - Rapport d'Inventaire")
    exporter.add_subtitle("Rapport d'inventaire complet")
    
    headers = ["Nom", "Référence", "Stock Disponible", "Stock Total", "Seuil Alerte", "État", "Note"]
    header_row = exporter.add_headers(headers)
    
    for eq in equipements:
        status = "Actif" if eq.est_actif else "Inactif"
        exporter.add_row(
            [
                eq.nom,
                eq.reference,
                eq.stock_disponible,
                eq.stock_total,
                eq.seuil_alerte,
                status,
                eq.note or "—",
            ],
            status_col=6,  # Colonne "État"
            status_value=status
        )
    
    exporter.auto_resize_columns()
    exporter.add_autofilter(header_row)
    
    return exporter.get_bytes(), exporter.get_filename()


def export_ateliers_excel(ateliers):
    """Exporter les ateliers en Excel formaté."""
    exporter = ExcelExporter("Ateliers", "ateliers")
    exporter.add_title("🏫 Gestion des Ateliers - Rapport Complet")
    exporter.add_subtitle("Liste de tous les ateliers")
    
    headers = ["Titre", "Date Début", "Date Fin", "Salle", "Tuteur", "Niveau Cible", "Créateur", "État"]
    header_row = exporter.add_headers(headers)
    
    for atelier in ateliers:
        statut = atelier.statut_dynamique
        exporter.add_row(
            [
                atelier.titre,
                atelier.date_debut.strftime("%d/%m/%Y %H:%M"),
                atelier.date_fin.strftime("%d/%m/%Y %H:%M"),
                atelier.salle,
                f"{atelier.tuteur.first_name} {atelier.tuteur.last_name}",
                atelier.niveau_cible,
                f"{atelier.createur.first_name} {atelier.createur.last_name}" if atelier.createur else "—",
                statut,
            ],
            status_col=8,  # Colonne "État"
            status_value=statut
        )
    
    exporter.auto_resize_columns()
    exporter.add_autofilter(header_row)
    
    return exporter.get_bytes(), exporter.get_filename()


def export_demandes_excel(demandes):
    """Exporter les demandes de matériel en Excel formaté."""
    exporter = ExcelExporter("Demandes", "demandes_materiel")
    exporter.add_title("🚚 Demandes de Matériel - Rapport Complet")
    exporter.add_subtitle("État des demandes de matériel")
    
    headers = ["Formateur", "Équipement", "Atelier Cible", "Quantité", "Statut", "Date Création"]
    header_row = exporter.add_headers(headers)
    
    for demande in demandes:
        statut_display = {
            "PENDING": "En attente",
            "APPROVED": "Approuvée",
            "REJECTED": "Rejetée",
            "RETURNED": "Retournée",
        }.get(demande.statut, demande.statut)
        
        exporter.add_row(
            [
                f"{demande.formateur.first_name} {demande.formateur.last_name}",
                demande.equipement.nom if demande.equipement else "—",
                demande.atelier_cible.titre if demande.atelier_cible else "—",
                demande.quantite,
                statut_display,
                demande.date_creation.strftime("%d/%m/%Y %H:%M"),
            ],
            status_col=5,  # Colonne "Statut"
            status_value=statut_display
        )
    
    exporter.auto_resize_columns()
    exporter.add_autofilter(header_row)
    
    return exporter.get_bytes(), exporter.get_filename()


def export_tickets_excel(tickets):
    """Exporter les tickets en Excel formaté."""
    exporter = ExcelExporter("Tickets", "tickets")
    exporter.add_title("🎫 Support & Tickets - Rapport Complet")
    exporter.add_subtitle("État des tickets de maintenance")
    
    headers = ["Titre", "Description", "Équipement", "Atelier", "Statut", "Ouvert par"]
    header_row = exporter.add_headers(headers)
    
    for ticket in tickets:
        exporter.add_row(
            [
                ticket.titre,
                ticket.description.replace("\n", " ")[:100],
                ticket.equipement.nom if ticket.equipement else "—",
                ticket.atelier.titre if ticket.atelier else "—",
                ticket.statut,
                f"{ticket.ouvert_par.first_name} {ticket.ouvert_par.last_name}",
            ],
            status_col=5,  # Colonne "Statut"
            status_value=ticket.statut
        )
    
    exporter.auto_resize_columns()
    exporter.add_autofilter(header_row)
    
    return exporter.get_bytes(), exporter.get_filename()
